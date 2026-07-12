"""
fetch_crypto_prices.py
----------------------
Fetches top cryptocurrencies by market capitalization from the free
CoinGecko API and saves the results to an Excel file (crypto_prices.xlsx).

Excel output includes:
  - Sheet "Crypto Prices": full list with row numbers and collection timestamp
  - Sheet "Summary": Top 10 by Market Cap and Top 10 by 24h price change

No API key is required for the public endpoint used here.
"""

from __future__ import annotations

import time
from datetime import datetime, timezone
from typing import Any

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# =============================================================================
# CONFIGURATION — change these values as needed
# =============================================================================

# How many coins to fetch (e.g. 20, 50, 100). Ordered by market cap (largest first).
TOP_N = 50

# Output Excel file name (created in the same folder as this script)
OUTPUT_FILE = "crypto_prices.xlsx"

# CoinGecko free API base URL
API_URL = "https://api.coingecko.com/api/v3/coins/markets"

# Max coins per API page (CoinGecko allows up to 250)
PER_PAGE = 250

# Seconds to wait between paginated requests (helps avoid rate limits)
REQUEST_DELAY_SECONDS = 1.5

# HTTP request timeout in seconds
REQUEST_TIMEOUT = 30

# How many times to retry a failed request (rate limit / temporary errors)
MAX_RETRIES = 3

# Extra wait (seconds) when API returns HTTP 429 (Too Many Requests)
RATE_LIMIT_WAIT_SECONDS = 30

# How many coins to show in each Summary block
SUMMARY_TOP_N = 10

# Column names used across the workbook (single source of truth)
COL_RANK = "#"
COL_NAME = "Name"
COL_SYMBOL = "Symbol"
COL_PRICE = "Current Price (USD)"
COL_MARKET_CAP = "Market Cap"
COL_CHANGE_24H = "24h Price Change (%)"
COL_VOLUME = "Total Volume (24h)"

DATA_COLUMNS = [
    COL_RANK,
    COL_NAME,
    COL_SYMBOL,
    COL_PRICE,
    COL_MARKET_CAP,
    COL_CHANGE_24H,
    COL_VOLUME,
]


# =============================================================================
# HELPERS
# =============================================================================

def utc_now_label() -> str:
    """Return a human-readable UTC timestamp, e.g. '2026-07-12 10:25 UTC'."""
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def safe_float(value: Any) -> float | None:
    """
    Convert a value to float safely.

    Returns None for missing / invalid values so Excel stays clean
    instead of crashing on bad API data.
    """
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def safe_str(value: Any, default: str = "") -> str:
    """Convert a value to a stripped string; empty if missing."""
    if value is None:
        return default
    return str(value).strip() or default


# =============================================================================
# DATA FETCHING
# =============================================================================

def _request_with_retries(params: dict[str, Any]) -> list[dict]:
    """
    Perform a GET request with basic retry logic.

    Retries on network errors, timeouts, HTTP 429 (rate limit),
    and temporary 5xx server errors.
    """
    last_error: Exception | None = None

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = requests.get(API_URL, params=params, timeout=REQUEST_TIMEOUT)

            # Rate limited — wait longer, then retry
            if response.status_code == 429:
                wait = RATE_LIMIT_WAIT_SECONDS
                print(
                    f"  Rate limit (HTTP 429). "
                    f"Waiting {wait}s before retry {attempt}/{MAX_RETRIES}..."
                )
                time.sleep(wait)
                continue

            # Temporary server issues
            if response.status_code >= 500:
                print(
                    f"  Server error HTTP {response.status_code}. "
                    f"Retry {attempt}/{MAX_RETRIES}..."
                )
                time.sleep(REQUEST_DELAY_SECONDS * attempt)
                continue

            response.raise_for_status()

            data = response.json()
            if not isinstance(data, list):
                raise ValueError(
                    f"Unexpected API response type: {type(data).__name__} "
                    "(expected a list of coins)."
                )
            return data

        except (requests.Timeout, requests.ConnectionError) as exc:
            last_error = exc
            print(
                f"  Network issue ({exc}). "
                f"Retry {attempt}/{MAX_RETRIES}..."
            )
            time.sleep(REQUEST_DELAY_SECONDS * attempt)

        except (requests.RequestException, ValueError, TypeError) as exc:
            # Non-retryable or last-attempt failures bubble up
            last_error = exc
            if attempt >= MAX_RETRIES:
                break
            print(f"  Request failed ({exc}). Retry {attempt}/{MAX_RETRIES}...")
            time.sleep(REQUEST_DELAY_SECONDS * attempt)

    raise RuntimeError(
        f"Failed to fetch data from CoinGecko after {MAX_RETRIES} attempts."
    ) from last_error


def fetch_top_coins(top_n: int) -> list[dict]:
    """
    Fetch the top `top_n` cryptocurrencies by market cap from CoinGecko.

    Uses the /coins/markets endpoint (paginated when top_n > PER_PAGE).

    Args:
        top_n: Number of coins to retrieve (must be >= 1).

    Returns:
        A list of raw coin dictionaries from the API.
    """
    if not isinstance(top_n, int) or top_n < 1:
        raise ValueError("TOP_N must be an integer >= 1.")

    all_coins: list[dict] = []
    page = 1
    remaining = top_n

    print(f"Fetching top {top_n} cryptocurrencies from CoinGecko...")

    while remaining > 0:
        per_page = min(remaining, PER_PAGE)
        params = {
            "vs_currency": "usd",
            "order": "market_cap_desc",
            "per_page": per_page,
            "page": page,
            "sparkline": "false",
            "price_change_percentage": "24h",
        }

        print(f"  Requesting page {page} ({per_page} coins)...")
        coins = _request_with_retries(params)

        if not coins:
            print("  No more coins returned by the API.")
            break

        all_coins.extend(coins)
        remaining -= len(coins)
        page += 1

        # Pause between pages to reduce chance of hitting rate limits
        if remaining > 0:
            print(f"  Waiting {REQUEST_DELAY_SECONDS}s before next request...")
            time.sleep(REQUEST_DELAY_SECONDS)

    return all_coins[:top_n]


def coins_to_dataframe(coins: list[dict]) -> pd.DataFrame:
    """
    Convert raw CoinGecko market data into a clean DataFrame.

    Adds a "#" rank column (1-based) at the front.
    Missing numeric fields become None instead of crashing.
    """
    rows: list[dict[str, Any]] = []

    for index, coin in enumerate(coins, start=1):
        if not isinstance(coin, dict):
            # Skip unexpected items so one bad record does not break the export
            print(f"  Warning: skipping invalid coin entry at position {index}.")
            continue

        rows.append(
            {
                COL_RANK: index,
                COL_NAME: safe_str(coin.get("name"), default="Unknown"),
                COL_SYMBOL: safe_str(coin.get("symbol")).upper(),
                COL_PRICE: safe_float(coin.get("current_price")),
                COL_MARKET_CAP: safe_float(coin.get("market_cap")),
                COL_CHANGE_24H: safe_float(coin.get("price_change_percentage_24h")),
                COL_VOLUME: safe_float(coin.get("total_volume")),
            }
        )

    df = pd.DataFrame(rows, columns=DATA_COLUMNS)

    # Re-number after skipping any bad rows so "#" stays continuous
    if not df.empty:
        df[COL_RANK] = range(1, len(df) + 1)

    return df


# =============================================================================
# SUMMARY DATA
# =============================================================================

def build_top_by_market_cap(df: pd.DataFrame, n: int = SUMMARY_TOP_N) -> pd.DataFrame:
    """Return top N coins by Market Cap (descending)."""
    if df.empty:
        return df.copy()

    result = (
        df.dropna(subset=[COL_MARKET_CAP])
        .sort_values(COL_MARKET_CAP, ascending=False)
        .head(n)
        .copy()
    )
    result[COL_RANK] = range(1, len(result) + 1)
    return result.reset_index(drop=True)


def build_top_by_24h_change(df: pd.DataFrame, n: int = SUMMARY_TOP_N) -> pd.DataFrame:
    """
    Return top N coins by 24h price change (highest / best performers first).

    Coins with missing 24h change values are excluded.
    """
    if df.empty:
        return df.copy()

    result = (
        df.dropna(subset=[COL_CHANGE_24H])
        .sort_values(COL_CHANGE_24H, ascending=False)
        .head(n)
        .copy()
    )
    result[COL_RANK] = range(1, len(result) + 1)
    return result.reset_index(drop=True)


# =============================================================================
# EXCEL STYLING HELPERS
# =============================================================================

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=12, color="1F4E79")
SECTION_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
BODY_FONT = Font(name="Arial", size=10)
META_FONT = Font(name="Arial", italic=True, size=10, color="595959")


def _style_header_row(ws, row: int, start_col: int, end_col: int) -> None:
    """Apply header style to a row of column titles."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_dataframe(
    ws,
    df: pd.DataFrame,
    start_row: int,
    start_col: int = 1,
) -> int:
    """
    Write a DataFrame (with header) starting at start_row/start_col.

    Returns the next empty row index after the written block.
    """
    if df.empty:
        ws.cell(row=start_row, column=start_col, value="(no data)").font = BODY_FONT
        return start_row + 1

    # Header
    for col_offset, column_name in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + col_offset, value=column_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Body rows
    for r_idx, row in enumerate(df.itertuples(index=False), start=1):
        excel_row = start_row + r_idx
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=excel_row, column=start_col + c_idx, value=value)
            cell.font = BODY_FONT

            # Apply number formats by column name
            col_name = df.columns[c_idx]
            if col_name == COL_PRICE:
                # Price: always 2 decimal places
                cell.number_format = "$#,##0.00"
            elif col_name in (COL_MARKET_CAP, COL_VOLUME):
                # Market Cap & Volume: thousand separators, whole dollars
                cell.number_format = "$#,##0"
            elif col_name == COL_CHANGE_24H:
                cell.number_format = "0.00"
            elif col_name == COL_RANK:
                cell.alignment = Alignment(horizontal="center")

    return start_row + len(df) + 1


def _set_column_widths(ws, widths: dict[str, float]) -> None:
    """Set column widths from a map like {'A': 8, 'B': 20, ...}."""
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def save_to_excel(df: pd.DataFrame, filepath: str, collected_at: str) -> None:
    """
    Save crypto data to Excel with two sheets:

    1) Crypto Prices — full table + collection timestamp in A1
    2) Summary — Top 10 by Market Cap and Top 10 by 24h Change
    """
    wb = Workbook()

    # -------------------------------------------------------------------------
    # Sheet 1: Crypto Prices
    # -------------------------------------------------------------------------
    ws_main = wb.active
    ws_main.title = "Crypto Prices"

    # Row 1: collection timestamp (requirement: date/time of data collection)
    ws_main["A1"] = f"Data collected: {collected_at}"
    ws_main["A1"].font = META_FONT
    ws_main.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DATA_COLUMNS))

    # Row 2: blank spacer
    # Row 3: headers + data
    data_start_row = 3
    _write_dataframe(ws_main, df, start_row=data_start_row)

    _set_column_widths(
        ws_main,
        {
            "A": 6,   # #
            "B": 22,  # Name
            "C": 12,  # Symbol
            "D": 20,  # Price
            "E": 18,  # Market Cap
            "F": 20,  # 24h Change
            "G": 18,  # Volume
        },
    )
    ws_main.freeze_panes = "A4"  # freeze timestamp + header
    ws_main.row_dimensions[data_start_row].height = 30

    # -------------------------------------------------------------------------
    # Sheet 2: Summary
    # -------------------------------------------------------------------------
    ws_sum = wb.create_sheet("Summary")

    ws_sum["A1"] = f"Data collected: {collected_at}"
    ws_sum["A1"].font = META_FONT
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DATA_COLUMNS))

    # --- Top 10 by Market Cap ---
    ws_sum["A3"] = f"Top {SUMMARY_TOP_N} by Market Cap"
    ws_sum["A3"].font = SECTION_FONT
    ws_sum["A3"].fill = SECTION_FILL
    for col in range(1, len(DATA_COLUMNS) + 1):
        ws_sum.cell(row=3, column=col).fill = SECTION_FILL
        ws_sum.cell(row=3, column=col).font = SECTION_FONT
    ws_sum.merge_cells(start_row=3, start_column=1, end_row=3, end_column=len(DATA_COLUMNS))

    top_mcap = build_top_by_market_cap(df, SUMMARY_TOP_N)
    next_row = _write_dataframe(ws_sum, top_mcap, start_row=4)

    # --- Top 10 by 24h Change ---
    section_row = next_row + 1  # one blank row between blocks
    ws_sum.cell(row=section_row, column=1, value=f"Top {SUMMARY_TOP_N} by 24h Price Change (%)")
    for col in range(1, len(DATA_COLUMNS) + 1):
        cell = ws_sum.cell(row=section_row, column=col)
        cell.fill = SECTION_FILL
        cell.font = SECTION_FONT
    ws_sum.merge_cells(
        start_row=section_row,
        start_column=1,
        end_row=section_row,
        end_column=len(DATA_COLUMNS),
    )

    top_change = build_top_by_24h_change(df, SUMMARY_TOP_N)
    _write_dataframe(ws_sum, top_change, start_row=section_row + 1)

    _set_column_widths(
        ws_sum,
        {
            "A": 6,
            "B": 22,
            "C": 12,
            "D": 20,
            "E": 18,
            "F": 20,
            "G": 18,
        },
    )

    # Save with a clear error if the file is locked (e.g. open in Excel)
    try:
        wb.save(filepath)
    except PermissionError as exc:
        raise PermissionError(
            f"Cannot write '{filepath}'. Close the file if it is open in Excel "
            "and try again."
        ) from exc

    print(f"Saved {len(df)} coins to '{filepath}' (sheets: Crypto Prices, Summary).")


# =============================================================================
# MAIN
# =============================================================================

def main() -> None:
    """Fetch crypto data and write the Excel workbook."""
    collected_at = utc_now_label()

    try:
        coins = fetch_top_coins(TOP_N)
    except ValueError as exc:
        print(f"Configuration error: {exc}")
        raise SystemExit(1) from exc
    except RuntimeError as exc:
        print(f"Failed to download data: {exc}")
        print(
            "Tips: check your internet connection, wait a minute if rate-limited, "
            "or increase REQUEST_DELAY_SECONDS / RATE_LIMIT_WAIT_SECONDS."
        )
        raise SystemExit(1) from exc
    except requests.RequestException as exc:
        print(f"Network error while contacting CoinGecko: {exc}")
        raise SystemExit(1) from exc

    if not coins:
        print("No data received from the API. Nothing to save.")
        raise SystemExit(1)

    df = coins_to_dataframe(coins)
    if df.empty:
        print("All coin records were invalid. Nothing to save.")
        raise SystemExit(1)

    # Terminal preview
    print("\nPreview (first 5 rows):")
    print(df.head().to_string(index=False))
    print()

    try:
        save_to_excel(df, OUTPUT_FILE, collected_at=collected_at)
    except PermissionError as exc:
        print(exc)
        raise SystemExit(1) from exc
    except OSError as exc:
        print(f"Could not write Excel file: {exc}")
        raise SystemExit(1) from exc

    print(f"Done. Data collected: {collected_at}")


if __name__ == "__main__":
    main()
