"""
Monthly sales report automation — polished summary workbook.

Usage:
    python create_sales_summary.py
    python create_sales_summary.py path/to/sales_report.xlsx
    python create_sales_summary.py path/to/sales_report.xlsx -o path/to/sales_summary.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Column aliases — script works when headers vary slightly
# ---------------------------------------------------------------------------
COLUMN_ALIASES = {
    "date": ["date", "order date", "sale date", "transaction date"],
    "region": ["region", "area", "territory", "market"],
    "category": ["product category", "category", "product_category", "cat"],
    "product": ["product name", "product", "item", "sku name", "product_name"],
    "quantity": ["quantity", "qty", "units", "units sold"],
    "unit_price": ["unit price", "price", "unit_price", "unitprice"],
    "revenue": [
        "revenue",
        "sales",
        "amount",
        "total",
        "total sales",
        "sales amount",
    ],
}

# ---------------------------------------------------------------------------
# Visual theme
# ---------------------------------------------------------------------------
CURRENCY = '$#,##0.00'
PERCENT = "0.0%"
NUMBER = "#,##0"
FONT = "Calibri"

NAVY = "1B4F72"
BLUE = "2E86AB"
LIGHT_BLUE = "D6EAF8"
SOFT_GRAY = "F4F6F7"
ALT_ROW = "EBF5FB"
WHITE = "FFFFFF"
DARK = "1C2833"
MUTED = "5D6D7E"
TOTAL_BG = "D5D8DC"
POS_BG = "C6EFCE"
POS_FG = "006100"
NEG_BG = "FFC7CE"
NEG_FG = "9C0006"
KPI_BG = "1B4F72"
KPI_ACCENT = "148F77"
GOLD = "B7950B"

THIN = Border(
    left=Side(style="thin", color="BFC9CA"),
    right=Side(style="thin", color="BFC9CA"),
    top=Side(style="thin", color="BFC9CA"),
    bottom=Side(style="thin", color="BFC9CA"),
)
MED_BOTTOM = Border(bottom=Side(style="medium", color=NAVY))

CHART_COLORS = ["2E86AB", "148F77", "E67E22", "8E44AD", "C0392B", "16A085", "F39C12"]


def _num(value: Any) -> float:
    """Coerce pandas/numpy scalars to plain Python float."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    return float(value)


def _normalize(name: str) -> str:
    return " ".join(str(name).strip().lower().split())


def resolve_columns(df: pd.DataFrame) -> dict[str, str]:
    normalized = {_normalize(c): c for c in df.columns}
    resolved: dict[str, str] = {}
    for logical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                resolved[logical] = normalized[alias]
                break
    required = ["region", "category", "product", "revenue"]
    missing = [r for r in required if r not in resolved]
    if missing:
        raise ValueError(
            "Missing required columns for: "
            + ", ".join(missing)
            + f". Found columns: {list(df.columns)}"
        )
    return resolved


def load_sales(path: Path) -> pd.DataFrame:
    xl = pd.ExcelFile(path)
    sheet = "Sales Data" if "Sales Data" in xl.sheet_names else xl.sheet_names[0]
    df = pd.read_excel(path, sheet_name=sheet).dropna(how="all")
    cols = resolve_columns(df)

    out = pd.DataFrame(
        {
            "Region": df[cols["region"]].astype(str).str.strip(),
            "Product Category": df[cols["category"]].astype(str).str.strip(),
            "Product Name": df[cols["product"]].astype(str).str.strip(),
            "Revenue": pd.to_numeric(df[cols["revenue"]], errors="coerce"),
        }
    )
    out["Date"] = (
        pd.to_datetime(df[cols["date"]], errors="coerce") if "date" in cols else pd.NaT
    )
    out["Quantity"] = (
        pd.to_numeric(df[cols["quantity"]], errors="coerce") if "quantity" in cols else pd.NA
    )
    out = out.dropna(subset=["Revenue"])

    if out["Date"].notna().any():
        out["MonthPeriod"] = out["Date"].dt.to_period("M")
        out["MonthLabel"] = out["MonthPeriod"].dt.strftime("%B %Y")
    else:
        out["MonthPeriod"] = pd.PeriodIndex(
            [pd.Period("1970-01", freq="M")] * len(out), freq="M"
        )
        out["MonthLabel"] = "Unknown"
    return out


def build_summaries(df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    total = float(df["Revenue"].sum())

    by_region = (
        df.groupby("Region", as_index=False)["Revenue"]
        .sum()
        .sort_values("Revenue", ascending=False)
        .reset_index(drop=True)
    )
    by_region["Share %"] = by_region["Revenue"] / total
    by_region.columns = ["Region", "Total Sales", "Share %"]

    by_category = (
        df.groupby("Product Category", as_index=False)["Revenue"]
        .sum()
        .sort_values("Revenue", ascending=False)
        .reset_index(drop=True)
    )
    by_category["Share %"] = by_category["Revenue"] / total
    by_category.columns = ["Product Category", "Total Sales", "Share %"]

    products = (
        df.groupby("Product Name", as_index=False)
        .agg(Quantity=("Quantity", "sum"), Revenue=("Revenue", "sum"))
        .sort_values("Revenue", ascending=False)
        .head(10)
        .reset_index(drop=True)
    )
    products.insert(0, "Rank", range(1, len(products) + 1))
    products["Share %"] = products["Revenue"] / total
    products.columns = ["Rank", "Product Name", "Quantity Sold", "Total Sales", "Share %"]

    monthly = (
        df.groupby(["MonthPeriod", "MonthLabel"], as_index=False)["Revenue"]
        .sum()
        .sort_values("MonthPeriod")
        .reset_index(drop=True)
    )
    monthly["Previous Sales"] = monthly["Revenue"].shift(1)
    monthly["Growth %"] = (monthly["Revenue"] - monthly["Previous Sales"]) / monthly[
        "Previous Sales"
    ]
    monthly = monthly.rename(columns={"MonthLabel": "Month", "Revenue": "Total Sales"})
    monthly = monthly[["Month", "Total Sales", "Previous Sales", "Growth %"]]

    return {
        "by_region": by_region,
        "by_category": by_category,
        "top_products": products,
        "monthly": monthly,
        "total": total,
    }


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
def font(size=11, bold=False, color=DARK, italic=False, name=FONT) -> Font:
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def set_cell(
    ws,
    row: int,
    col: int,
    value: Any,
    *,
    bold: bool = False,
    size: int = 11,
    color: str = DARK,
    bg: str | None = None,
    num_format: str | None = None,
    align: str = "left",
    border: bool = True,
) -> Any:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font(size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg:
        cell.fill = fill(bg)
    if num_format and isinstance(value, (int, float)) and not isinstance(value, bool):
        cell.number_format = num_format
    if border:
        cell.border = THIN
    return cell


def paint_header(ws, row: int, cols: int, bg: str = NAVY) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font(size=11, bold=True, color=WHITE)
        cell.fill = fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.row_dimensions[row].height = 22


def paint_title(ws, row: int, col: int, text: str, size: int = 16) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font(size=size, bold=True, color=NAVY)
    cell.alignment = Alignment(vertical="center")


def paint_subtitle(ws, row: int, col: int, text: str) -> None:
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font(size=10, color=MUTED, italic=True)


def zebra(ws, start: int, end: int, cols: int) -> None:
    for r in range(start, end + 1):
        if (r - start) % 2 == 1:
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                if not cell.fill or cell.fill.fgColor is None or cell.fill.fgColor.rgb in (
                    "00000000",
                    "0",
                    None,
                ):
                    cell.fill = fill(ALT_ROW)


def total_row(ws, row: int, label: str, values: dict[int, tuple[Any, str | None]]) -> None:
    """values: {col: (value, number_format)}"""
    for c in range(1, max(values) + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill(TOTAL_BG)
        cell.font = font(bold=True, size=11)
        cell.border = THIN
        cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center")
    ws.cell(row=row, column=1, value=label)
    for c, (val, fmt) in values.items():
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = font(bold=True, size=11)
        cell.fill = fill(TOTAL_BG)
        cell.border = THIN
        if fmt and isinstance(val, (int, float)):
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right", vertical="center")


def set_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def apply_growth_style(cell, growth: float | None) -> None:
    if growth is None or (isinstance(growth, float) and pd.isna(growth)):
        cell.value = "—"
        cell.font = font(size=11, color=MUTED)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return
    cell.value = float(growth)
    cell.number_format = PERCENT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if growth > 0:
        cell.fill = fill(POS_BG)
        cell.font = font(size=11, bold=True, color=POS_FG)
    elif growth < 0:
        cell.fill = fill(NEG_BG)
        cell.font = font(size=11, bold=True, color=NEG_FG)
    else:
        cell.font = font(size=11, color=MUTED)


def color_chart_series(chart, n_points: int) -> None:
    """Assign distinct solid colors to points of the first series (bar charts)."""
    if not chart.series:
        return
    series = chart.series[0]
    pts = []
    for i in range(n_points):
        pt = DataPoint(idx=i)
        color = CHART_COLORS[i % len(CHART_COLORS)]
        pt.graphicalProperties.solidFill = color
        pts.append(pt)
    series.data_points = pts


def make_bar_chart(
    ws,
    *,
    title: str,
    data_col: int,
    cat_col: int,
    header_row: int,
    data_start: int,
    data_end: int,
    anchor: str,
    width: float = 15,
    height: float = 10,
    y_title: str = "Total Sales ($)",
    style: int = 10,
) -> BarChart:
    chart = BarChart()
    chart.type = "col"
    chart.style = style
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = None
    chart.legend = None
    data = Reference(ws, min_col=data_col, min_row=header_row, max_row=data_end)
    cats = Reference(ws, min_col=cat_col, min_row=data_start, max_row=data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = width
    chart.height = height
    color_chart_series(chart, data_end - data_start + 1)
    ws.add_chart(chart, anchor)
    return chart


def make_line_chart(
    ws,
    *,
    title: str,
    data_col: int,
    cat_col: int,
    header_row: int,
    data_start: int,
    data_end: int,
    anchor: str,
    width: float = 14,
    height: float = 9,
) -> LineChart:
    chart = LineChart()
    chart.style = 10
    chart.title = title
    chart.y_axis.title = "Total Sales ($)"
    chart.legend = None
    data = Reference(ws, min_col=data_col, min_row=header_row, max_row=data_end)
    cats = Reference(ws, min_col=cat_col, min_row=data_start, max_row=data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = width
    chart.height = height
    if chart.series:
        chart.series[0].graphicalProperties.line.solidFill = BLUE
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.size = 8
        chart.series[0].marker.graphicalProperties.solidFill = BLUE
    ws.add_chart(chart, anchor)
    return chart


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def build_overview(wb: Workbook, df: pd.DataFrame, s: dict) -> None:
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_view.showGridLines = False

    total = s["total"]
    region = s["by_region"]
    category = s["by_category"]
    monthly = s["monthly"]
    n_tx = len(df)
    n_regions = int(df["Region"].nunique())
    n_cats = int(df["Product Category"].nunique())
    n_products = int(df["Product Name"].nunique())

    date_min, date_max = df["Date"].min(), df["Date"].max()
    period = ""
    if pd.notna(date_min) and pd.notna(date_max):
        period = f"{date_min.strftime('%d %b %Y')} — {date_max.strftime('%d %b %Y')}"

    # Banner
    ws.merge_cells("A1:J1")
    paint_title(ws, 1, 1, "SALES SUMMARY REPORT", size=20)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:J2")
    paint_subtitle(
        ws,
        2,
        1,
        f"Period: {period}   ·   {n_tx:,} transactions   ·   Auto-generated summary",
    )
    ws.row_dimensions[2].height = 18

    # Accent line
    for c in range(1, 11):
        ws.cell(row=3, column=c).border = MED_BOTTOM

    # KPI cards — row 5 labels, row 6 values
    kpis = [
        (1, 2, "TOTAL REVENUE", total, CURRENCY, KPI_BG),
        (3, 4, "TRANSACTIONS", n_tx, NUMBER, BLUE),
        (5, 6, "REGIONS", n_regions, NUMBER, KPI_ACCENT),
        (7, 8, "CATEGORIES", n_cats, NUMBER, GOLD),
        (9, 10, "PRODUCTS", n_products, NUMBER, "6C3483"),
    ]
    for c1, c2, label, value, fmt, bg in kpis:
        ws.merge_cells(start_row=5, start_column=c1, end_row=5, end_column=c2)
        ws.merge_cells(start_row=6, start_column=c1, end_row=6, end_column=c2)
        lbl = ws.cell(row=5, column=c1, value=label)
        lbl.font = font(size=9, bold=True, color=WHITE)
        lbl.fill = fill(bg)
        lbl.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(c1, c2 + 1):
            ws.cell(row=5, column=col).fill = fill(bg)
            ws.cell(row=5, column=col).border = THIN
            ws.cell(row=6, column=col).fill = fill(bg)
            ws.cell(row=6, column=col).border = THIN
        val = ws.cell(row=6, column=c1, value=value)
        val.font = font(size=16, bold=True, color=WHITE)
        val.fill = fill(bg)
        val.alignment = Alignment(horizontal="center", vertical="center")
        if isinstance(value, float) or (isinstance(value, (int, float)) and fmt == CURRENCY):
            val.number_format = fmt
        elif fmt == NUMBER:
            val.number_format = NUMBER
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 30

    # Latest month growth highlight
    last_growth = None
    last_month = None
    if len(monthly) >= 2:
        last = monthly.iloc[-1]
        last_month = last["Month"]
        g = last["Growth %"]
        if pd.notna(g):
            last_growth = float(g)

    ws.merge_cells("A8:C8")
    paint_title(ws, 8, 1, "Key insight", size=12)
    ws.merge_cells("A9:C9")
    if last_growth is not None and last_month:
        direction = "up" if last_growth >= 0 else "down"
        insight = f"{last_month}: sales {direction} {abs(last_growth):.1%} vs previous month"
        cell = ws.cell(row=9, column=1, value=insight)
        cell.font = font(
            size=12,
            bold=True,
            color=POS_FG if last_growth >= 0 else NEG_FG,
        )
        cell.fill = fill(POS_BG if last_growth >= 0 else NEG_BG)
        for c in range(1, 4):
            ws.cell(row=9, column=c).fill = fill(POS_BG if last_growth >= 0 else NEG_BG)
            ws.cell(row=9, column=c).border = THIN
    else:
        paint_subtitle(ws, 9, 1, "Not enough months to compute growth.")

    # ---- Region table (left) ----
    paint_title(ws, 11, 1, "Sales by Region", size=13)
    headers_r = ["Region", "Total Sales", "Share %"]
    for i, h in enumerate(headers_r, 1):
        set_cell(ws, 12, i, h, bold=True, color=WHITE, bg=NAVY, align="center")
    paint_header(ws, 12, 3)

    for i, row in region.iterrows():
        r = 13 + i
        bg = ALT_ROW if i % 2 else WHITE
        set_cell(ws, r, 1, str(row["Region"]), bg=bg)
        set_cell(ws, r, 2, _num(row["Total Sales"]), bg=bg, num_format=CURRENCY, align="right")
        set_cell(ws, r, 3, _num(row["Share %"]), bg=bg, num_format=PERCENT, align="center")

    end_r = 12 + len(region)
    total_row(
        ws,
        end_r + 1,
        "Total",
        {2: (total, CURRENCY), 3: (1.0, PERCENT)},
    )

    make_bar_chart(
        ws,
        title="Sales by Region",
        data_col=2,
        cat_col=1,
        header_row=12,
        data_start=13,
        data_end=end_r,
        anchor="E11",
        width=14,
        height=9,
    )

    # ---- Category table ----
    cat_title_row = end_r + 4
    paint_title(ws, cat_title_row, 1, "Sales by Product Category", size=13)
    for i, h in enumerate(["Product Category", "Total Sales", "Share %"], 1):
        set_cell(ws, cat_title_row + 1, i, h, bold=True, color=WHITE, bg=NAVY, align="center")
    paint_header(ws, cat_title_row + 1, 3)

    for i, row in category.iterrows():
        r = cat_title_row + 2 + i
        bg = ALT_ROW if i % 2 else WHITE
        set_cell(ws, r, 1, str(row["Product Category"]), bg=bg)
        set_cell(ws, r, 2, _num(row["Total Sales"]), bg=bg, num_format=CURRENCY, align="right")
        set_cell(ws, r, 3, _num(row["Share %"]), bg=bg, num_format=PERCENT, align="center")

    end_c = cat_title_row + 1 + len(category)
    total_row(
        ws,
        end_c + 1,
        "Total",
        {2: (total, CURRENCY), 3: (1.0, PERCENT)},
    )

    make_bar_chart(
        ws,
        title="Sales by Category",
        data_col=2,
        cat_col=1,
        header_row=cat_title_row + 1,
        data_start=cat_title_row + 2,
        data_end=end_c,
        anchor="E" + str(cat_title_row),
        width=14,
        height=9,
        style=11,
    )

    set_widths(ws, {"A": 22, "B": 15, "C": 12, "D": 3, "E": 12, "F": 12, "G": 12, "H": 12, "I": 12, "J": 12})
    ws.freeze_panes = "A5"
    ws.print_title_rows = "1:2"


def build_region_sheet(wb: Workbook, s: dict) -> None:
    ws = wb.create_sheet("Sales by Region")
    ws.sheet_view.showGridLines = False
    region = s["by_region"]
    total = s["total"]

    paint_title(ws, 1, 1, "Total Sales by Region", size=16)
    paint_subtitle(ws, 2, 1, "Breakdown of revenue across all regions")

    headers = ["Region", "Total Sales", "Share %"]
    for i, h in enumerate(headers, 1):
        set_cell(ws, 4, i, h, bold=True, color=WHITE, bg=NAVY, align="center")
    paint_header(ws, 4, 3)

    for i, row in region.iterrows():
        r = 5 + i
        bg = ALT_ROW if i % 2 else WHITE
        set_cell(ws, r, 1, str(row["Region"]), bg=bg)
        set_cell(ws, r, 2, _num(row["Total Sales"]), bg=bg, num_format=CURRENCY, align="right")
        set_cell(ws, r, 3, _num(row["Share %"]), bg=bg, num_format=PERCENT, align="center")

    end = 4 + len(region)
    total_row(ws, end + 1, "Total", {2: (total, CURRENCY), 3: (1.0, PERCENT)})

    make_bar_chart(
        ws,
        title="Sales by Region",
        data_col=2,
        cat_col=1,
        header_row=4,
        data_start=5,
        data_end=end,
        anchor="E4",
        width=16,
        height=11,
    )
    set_widths(ws, {"A": 20, "B": 16, "C": 12, "D": 3, "E": 14})
    ws.freeze_panes = "A5"


def build_category_sheet(wb: Workbook, s: dict) -> None:
    ws = wb.create_sheet("Sales by Category")
    ws.sheet_view.showGridLines = False
    category = s["by_category"]
    total = s["total"]

    paint_title(ws, 1, 1, "Total Sales by Product Category", size=16)
    paint_subtitle(ws, 2, 1, "Revenue by product category")

    for i, h in enumerate(["Product Category", "Total Sales", "Share %"], 1):
        set_cell(ws, 4, i, h, bold=True, color=WHITE, bg=NAVY, align="center")
    paint_header(ws, 4, 3)

    for i, row in category.iterrows():
        r = 5 + i
        bg = ALT_ROW if i % 2 else WHITE
        set_cell(ws, r, 1, str(row["Product Category"]), bg=bg)
        set_cell(ws, r, 2, _num(row["Total Sales"]), bg=bg, num_format=CURRENCY, align="right")
        set_cell(ws, r, 3, _num(row["Share %"]), bg=bg, num_format=PERCENT, align="center")

    end = 4 + len(category)
    total_row(ws, end + 1, "Total", {2: (total, CURRENCY), 3: (1.0, PERCENT)})

    make_bar_chart(
        ws,
        title="Sales by Category",
        data_col=2,
        cat_col=1,
        header_row=4,
        data_start=5,
        data_end=end,
        anchor="E4",
        width=16,
        height=11,
        style=12,
    )
    set_widths(ws, {"A": 22, "B": 16, "C": 12})
    ws.freeze_panes = "A5"


def build_top_products(wb: Workbook, s: dict) -> None:
    ws = wb.create_sheet("Top 10 Products")
    ws.sheet_view.showGridLines = False
    top = s["top_products"]

    paint_title(ws, 1, 1, "Top 10 Best-Selling Products", size=16)
    paint_subtitle(ws, 2, 1, "Ranked by total revenue")

    headers = ["Rank", "Product Name", "Quantity Sold", "Total Sales", "Share %"]
    for i, h in enumerate(headers, 1):
        set_cell(ws, 4, i, h, bold=True, color=WHITE, bg=NAVY, align="center")
    paint_header(ws, 4, 5)

    for i, row in top.iterrows():
        r = 5 + i
        bg = ALT_ROW if i % 2 else WHITE
        # Gold / silver / bronze for top 3
        if i == 0:
            bg = "FCF3CF"
        elif i == 1:
            bg = "E5E8E8"
        elif i == 2:
            bg = "F5CBA7"

        set_cell(ws, r, 1, int(row["Rank"]), bg=bg, align="center", bold=True)
        set_cell(ws, r, 2, str(row["Product Name"]), bg=bg)
        qty = row["Quantity Sold"]
        qty_val = _num(qty) if pd.notna(qty) else None
        if qty_val is not None:
            set_cell(ws, r, 3, qty_val, bg=bg, num_format=NUMBER, align="right")
        else:
            set_cell(ws, r, 3, "—", bg=bg, align="center", color=MUTED)
        set_cell(ws, r, 4, _num(row["Total Sales"]), bg=bg, num_format=CURRENCY, align="right")
        set_cell(ws, r, 5, _num(row["Share %"]), bg=bg, num_format=PERCENT, align="center")

    end = 4 + len(top)
    make_bar_chart(
        ws,
        title="Top 10 Products by Revenue",
        data_col=4,
        cat_col=2,
        header_row=4,
        data_start=5,
        data_end=end,
        anchor="G4",
        width=16,
        height=12,
        y_title="Revenue ($)",
        style=10,
    )
    set_widths(ws, {"A": 8, "B": 22, "C": 14, "D": 15, "E": 12, "F": 3, "G": 14})
    ws.freeze_panes = "A5"


def build_monthly_growth(wb: Workbook, s: dict) -> None:
    ws = wb.create_sheet("Monthly Growth")
    ws.sheet_view.showGridLines = False
    monthly = s["monthly"]

    paint_title(ws, 1, 1, "Monthly Growth vs Previous Month", size=16)
    paint_subtitle(
        ws,
        2,
        1,
        "Growth % = (Current − Previous) / Previous. First month has no prior period.",
    )

    headers = ["Month", "Total Sales", "Previous Sales", "Change ($)", "Growth %"]
    for i, h in enumerate(headers, 1):
        set_cell(ws, 4, i, h, bold=True, color=WHITE, bg=NAVY, align="center")
    paint_header(ws, 4, 5)

    for i, row in monthly.iterrows():
        r = 5 + i
        bg = ALT_ROW if i % 2 else WHITE
        set_cell(ws, r, 1, str(row["Month"]), bg=bg, bold=True)
        set_cell(ws, r, 2, _num(row["Total Sales"]), bg=bg, num_format=CURRENCY, align="right")

        prev = row["Previous Sales"]
        growth = row["Growth %"]

        if pd.isna(prev):
            set_cell(ws, r, 3, "—", bg=bg, align="center", color=MUTED)
            set_cell(ws, r, 4, "—", bg=bg, align="center", color=MUTED)
            gcell = set_cell(ws, r, 5, "—", bg=bg, align="center", color=MUTED)
            apply_growth_style(gcell, None)
        else:
            prev_f = _num(prev)
            sales_f = _num(row["Total Sales"])
            change = sales_f - prev_f
            set_cell(ws, r, 3, prev_f, bg=bg, num_format=CURRENCY, align="right")

            ch = set_cell(ws, r, 4, change, bg=bg, num_format=CURRENCY, align="right")
            if change > 0:
                ch.font = font(size=11, bold=True, color=POS_FG)
                ch.fill = fill(POS_BG)
            elif change < 0:
                ch.font = font(size=11, bold=True, color=NEG_FG)
                ch.fill = fill(NEG_BG)

            # Pre-calculated percentage value (visible immediately, no Excel recalc needed)
            gcell = ws.cell(row=r, column=5)
            gcell.border = THIN
            apply_growth_style(gcell, float(growth) if pd.notna(growth) else None)

        ws.row_dimensions[r].height = 20

    end = 4 + len(monthly)

    # Legend
    legend = end + 2
    set_cell(ws, legend, 1, "Legend", bold=True, border=False, size=11)
    pos = set_cell(ws, legend, 2, "Positive growth", bg=POS_BG, color=POS_FG, bold=True, align="center")
    neg = set_cell(ws, legend, 3, "Negative growth", bg=NEG_BG, color=NEG_FG, bold=True, align="center")

    # Line chart of monthly sales
    make_line_chart(
        ws,
        title="Monthly Sales Trend",
        data_col=2,
        cat_col=1,
        header_row=4,
        data_start=5,
        data_end=end,
        anchor="G4",
        width=14,
        height=9,
    )

    # Bar chart for growth % (helper column with numeric only for chart)
    # Put chart-friendly growth values in column K (hidden-ish side area)
    set_cell(ws, 4, 11, "Growth (chart)", bold=True, color=WHITE, bg=NAVY, align="center")
    for i, row in monthly.iterrows():
        r = 5 + i
        g = row["Growth %"]
        val = _num(g) if pd.notna(g) else 0.0
        cell = set_cell(ws, r, 11, val, num_format=PERCENT, align="center")
        if pd.notna(g):
            apply_growth_style(cell, float(g))
        else:
            cell.value = 0.0
            cell.number_format = PERCENT

    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Month-over-Month Growth %"
    chart.y_axis.title = "Growth %"
    chart.legend = None
    data = Reference(ws, min_col=11, min_row=4, max_row=end)
    cats = Reference(ws, min_col=1, min_row=5, max_row=end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 14
    chart.height = 9
    color_chart_series(chart, end - 5 + 1)
    ws.add_chart(chart, "G16")

    ws.column_dimensions["K"].hidden = True
    set_widths(ws, {"A": 16, "B": 15, "C": 16, "D": 14, "E": 12, "F": 3, "G": 14})
    ws.freeze_panes = "A5"


def create_workbook(df: pd.DataFrame, summaries: dict) -> Workbook:
    wb = Workbook()
    build_overview(wb, df, summaries)
    build_region_sheet(wb, summaries)
    build_category_sheet(wb, summaries)
    build_top_products(wb, summaries)
    build_monthly_growth(wb, summaries)
    return wb


def run(input_path: Path, output_path: Path) -> None:
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    df = load_sales(input_path)
    if df.empty:
        raise ValueError("No sales rows found after cleaning the input file.")

    summaries = build_summaries(df)
    wb = create_workbook(df, summaries)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    monthly = summaries["monthly"]
    print(f"Summary saved to: {output_path.resolve()}")
    print(f"  Transactions : {len(df):,}")
    print(f"  Total revenue: ${summaries['total']:,.2f}")
    print(f"  Regions      : {df['Region'].nunique()}")
    print(f"  Categories   : {df['Product Category'].nunique()}")
    print(f"  Months       : {df['MonthLabel'].nunique()}")
    print("  Growth %:")
    for _, row in monthly.iterrows():
        g = row["Growth %"]
        gtxt = f"{float(g):+.1%}" if pd.notna(g) else "n/a (first month)"
        print(f"    {row['Month']}: {gtxt}  (${_num(row['Total Sales']):,.2f})")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a formatted sales summary Excel report from a monthly sales file."
    )
    parser.add_argument(
        "input",
        nargs="?",
        default="sales_report.xlsx",
        help="Path to the source sales Excel file (default: sales_report.xlsx)",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="sales_summary.xlsx",
        help="Path for the summary output (default: sales_summary.xlsx)",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        run(Path(args.input), Path(args.output))
    except Exception as exc:  # noqa: BLE001
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
