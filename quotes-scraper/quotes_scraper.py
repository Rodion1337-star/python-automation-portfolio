# ============================================================
# Quotes Scraper — http://quotes.toscrape.com/
#
# Collects: quote text, author name, tags
# Saves to: quotes.xlsx
#
# Setup (once):
#   pip install requests beautifulsoup4 openpyxl
#
# Run:
#   python quotes_scraper.py        # 5 pages (default)
#   python quotes_scraper.py 10     # 10 pages
#
# See README.txt for a short how-to guide.
# ============================================================

# ------------------------------------------------------------
# 1. Imports
# ------------------------------------------------------------
import sys
import time
from pathlib import Path

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ------------------------------------------------------------
# 2. Settings (easy to change)
# ------------------------------------------------------------
BASE_URL = "http://quotes.toscrape.com"
DEFAULT_PAGES = 5
DELAY_SECONDS = 2.5
OUTPUT_FILE = "quotes.xlsx"

# Network resilience
REQUEST_TIMEOUT = 15          # seconds per request
MAX_RETRIES = 3               # retries for a failed page
RETRY_DELAY_SECONDS = 3       # wait before retry
MAX_CONSECUTIVE_FAILURES = 2  # stop if this many pages fail in a row


# ------------------------------------------------------------
# 3. Styles for Excel
# ------------------------------------------------------------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")


# ------------------------------------------------------------
# 4. Fetch HTML (with retries)
# ------------------------------------------------------------
def get_html(url: str, session: requests.Session) -> str | None:
    """
    Download HTML for a URL.
    Retries on temporary network/HTTP errors.
    Returns HTML text, or None if all attempts fail.
    """
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "en-US,en;q=0.9",
    }

    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = session.get(
                url,
                headers=headers,
                timeout=REQUEST_TIMEOUT,
            )
            response.raise_for_status()

            # Prefer server encoding; fall back safely
            if not response.encoding:
                response.encoding = response.apparent_encoding or "utf-8"

            text = response.text
            if not text or not text.strip():
                print(f"[Warning] Empty response from: {url}")
                return None
            return text

        except requests.exceptions.Timeout:
            print(f"[Error] Timeout ({attempt}/{MAX_RETRIES}): {url}")
        except requests.exceptions.HTTPError as error:
            status = error.response.status_code if error.response is not None else "?"
            # 404 often means "no more pages" — do not retry endlessly
            if status == 404:
                print(f"[Info] Page not found (404): {url}")
                return None
            print(f"[Error] HTTP {status} ({attempt}/{MAX_RETRIES}): {url}")
        except requests.exceptions.ConnectionError:
            print(f"[Error] Connection failed ({attempt}/{MAX_RETRIES}): {url}")
        except requests.exceptions.RequestException as error:
            print(f"[Error] Request failed ({attempt}/{MAX_RETRIES}): {error}")

        if attempt < MAX_RETRIES:
            print(f"[Info] Retrying in {RETRY_DELAY_SECONDS}s...")
            time.sleep(RETRY_DELAY_SECONDS)

    return None


# ------------------------------------------------------------
# 5. Parse one page
# ------------------------------------------------------------
def parse_quotes(html: str) -> list[dict]:
    """
    Extract quotes from one HTML page.

    Site structure (quotes.toscrape.com):
      div.quote
        span.text       → quote text
        small.author    → author
        div.tags a.tag  → tags

    Returns list of:
      {"quote_text": str, "author": str, "tags": list[str]}
    Skips incomplete blocks instead of crashing.
    """
    soup = BeautifulSoup(html, "html.parser")
    quotes: list[dict] = []

    for block in soup.select("div.quote"):
        try:
            text_tag = block.select_one("span.text")
            author_tag = block.select_one("small.author")

            quote_text = text_tag.get_text(strip=True) if text_tag else ""
            author = author_tag.get_text(strip=True) if author_tag else ""

            # Clean decorative quotes if present
            quote_text = quote_text.strip("“”\"'")

            tags = [
                tag.get_text(strip=True)
                for tag in block.select("div.tags a.tag")
                if tag.get_text(strip=True)
            ]

            # Skip empty / broken cards
            if not quote_text and not author:
                continue

            quotes.append(
                {
                    "quote_text": quote_text or "(no text)",
                    "author": author or "(unknown author)",
                    "tags": tags,
                }
            )
        except Exception as error:
            # One bad card must not stop the whole page
            print(f"[Warning] Skipped one quote block: {error}")
            continue

    return quotes


# ------------------------------------------------------------
# 6. Build page URL
# ------------------------------------------------------------
def build_page_url(page_number: int) -> str:
    """Build URL for page N, e.g. .../page/2/"""
    return f"{BASE_URL.rstrip('/')}/page/{page_number}/"


# ------------------------------------------------------------
# 7. Scrape several pages
# ------------------------------------------------------------
def scrape_pages(max_pages: int) -> list[dict]:
    """
    Visit pages 1..max_pages and collect all quotes.
    Stops early if pages end or too many failures in a row.
    """
    all_quotes: list[dict] = []
    consecutive_failures = 0

    with requests.Session() as session:
        for page_number in range(1, max_pages + 1):
            url = build_page_url(page_number)
            print(f"[Info] Page {page_number}/{max_pages}: {url}")

            html = get_html(url, session=session)
            if html is None:
                consecutive_failures += 1
                print(
                    f"[Warning] Page {page_number} failed "
                    f"({consecutive_failures} in a row)."
                )
                if consecutive_failures >= MAX_CONSECUTIVE_FAILURES:
                    print(
                        "[Warning] Too many failed pages in a row. "
                        "Stopping to avoid useless requests."
                    )
                    break
            else:
                consecutive_failures = 0
                try:
                    page_quotes = parse_quotes(html)
                except Exception as error:
                    print(f"[Error] Could not parse page {page_number}: {error}")
                    page_quotes = []

                if not page_quotes:
                    print(
                        f"[Warning] No quotes on page {page_number}. "
                        "Stopping (end of catalog or empty page)."
                    )
                    break

                all_quotes.extend(page_quotes)
                print(f"[Info] Found on this page: {len(page_quotes)}")
                print(f"[Info] Total collected: {len(all_quotes)}")

            # Polite delay before next request
            if page_number < max_pages:
                print(f"[Info] Waiting {DELAY_SECONDS}s...")
                time.sleep(DELAY_SECONDS)

    return all_quotes


# ------------------------------------------------------------
# 8. Excel helpers
# ------------------------------------------------------------
def style_header_row(sheet: Worksheet, column_count: int) -> None:
    """Apply header style to the first row."""
    sheet.row_dimensions[1].height = 22
    for col in range(1, column_count + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_data_cell(cell, row_index: int) -> None:
    """Apply body style; alternate row shading."""
    cell.alignment = CELL_ALIGN
    cell.border = THIN_BORDER
    if row_index % 2 == 0:
        cell.fill = ALT_ROW_FILL


def finish_sheet(sheet: Worksheet, last_row: int, last_col: int) -> None:
    """Freeze header, add filter, enable text wrap layout."""
    if last_row < 1 or last_col < 1:
        return
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"
    sheet.sheet_view.showGridLines = False


def resolve_output_path(filename: str) -> Path:
    """
    If the target file is locked (open in Excel), save under a new name.
    """
    path = Path(filename)
    if not path.exists():
        return path

    # Try opening for append to detect lock on Windows
    try:
        with open(path, "a"):
            pass
        return path
    except OSError:
        stamp = time.strftime("%Y%m%d_%H%M%S")
        alt = path.with_name(f"{path.stem}_{stamp}{path.suffix}")
        print(f"[Warning] '{path.name}' is locked. Saving as '{alt.name}' instead.")
        return alt


# ------------------------------------------------------------
# 9. Save to Excel
# ------------------------------------------------------------
def save_to_excel(data: list[dict], filename: str = OUTPUT_FILE) -> Path | None:
    """
    Save results into a formatted workbook with two sheets:

    1) All Quotes
       One row = one quote
       Columns: # | Quote Text | Author Name | Tag 1 | Tag 2 | ...

    2) Tags Lookup  (easier filtering)
       One row = one tag
       Columns: Tag Name | Author Name | Quote Text | Quote #
       Tip: filter by "Tag Name" to find all quotes with that tag.
    """
    if not data:
        print("[Warning] No data to save.")
        return None

    output_path = resolve_output_path(filename)

    try:
        workbook = Workbook()

        # ========== Sheet 1: All Quotes ==========
        sheet_main = workbook.active
        sheet_main.title = "All Quotes"

        max_tags = max((len(item.get("tags") or []) for item in data), default=0)
        tag_headers = [f"Tag {i}" for i in range(1, max_tags + 1)]
        main_headers = ["#", "Quote Text", "Author Name", *tag_headers]

        for col, name in enumerate(main_headers, start=1):
            sheet_main.cell(row=1, column=col, value=name)
        style_header_row(sheet_main, len(main_headers))

        for row_index, item in enumerate(data, start=2):
            quote_no = row_index - 1
            values = [
                quote_no,
                item.get("quote_text", ""),
                item.get("author", ""),
            ]
            tags = item.get("tags") or []
            values.extend(tags[i] if i < len(tags) else "" for i in range(max_tags))

            for col, value in enumerate(values, start=1):
                cell = sheet_main.cell(row=row_index, column=col, value=value)
                style_data_cell(cell, row_index)

        sheet_main.column_dimensions["A"].width = 6
        sheet_main.column_dimensions["B"].width = 70
        sheet_main.column_dimensions["C"].width = 22
        for col in range(4, 4 + max_tags):
            sheet_main.column_dimensions[get_column_letter(col)].width = 16

        last_main_row = 1 + len(data)
        finish_sheet(sheet_main, last_main_row, len(main_headers))

        # ========== Sheet 2: Tags Lookup ==========
        # Purpose: filter/search by a single tag easily
        sheet_tags = workbook.create_sheet("Tags Lookup")

        # Tag first — most useful column for filtering
        tag_headers_full = [
            "Tag Name",
            "Author Name",
            "Quote Text",
            "Quote #",
        ]
        for col, name in enumerate(tag_headers_full, start=1):
            sheet_tags.cell(row=1, column=col, value=name)
        style_header_row(sheet_tags, len(tag_headers_full))

        # Optional note row under header area is avoided (breaks filter).
        # Sheet tab name + clear headers explain the purpose.

        row_index = 2
        tag_rows = 0
        for quote_no, item in enumerate(data, start=1):
            tags = item.get("tags") or []
            if not tags:
                # Keep quote visible even without tags
                tags = ["(no tags)"]

            for tag in tags:
                values = [
                    tag,
                    item.get("author", ""),
                    item.get("quote_text", ""),
                    quote_no,
                ]
                for col, value in enumerate(values, start=1):
                    cell = sheet_tags.cell(row=row_index, column=col, value=value)
                    style_data_cell(cell, row_index)
                row_index += 1
                tag_rows += 1

        sheet_tags.column_dimensions["A"].width = 22
        sheet_tags.column_dimensions["B"].width = 22
        sheet_tags.column_dimensions["C"].width = 70
        sheet_tags.column_dimensions["D"].width = 10

        last_tag_row = max(1, row_index - 1)
        finish_sheet(sheet_tags, last_tag_row, len(tag_headers_full))

        # ========== Sheet 3: How to use (short help) ==========
        sheet_help = workbook.create_sheet("How to use")
        help_lines = [
            "How to use this file",
            "",
            "Sheet «All Quotes»",
            "  • One row = one quote.",
            "  • Columns: #, Quote Text, Author Name, Tag 1, Tag 2, ...",
            "  • Use filters on the header row to search by author or tag column.",
            "",
            "Sheet «Tags Lookup»",
            "  • One row = one tag linked to a quote.",
            "  • Use the filter on «Tag Name» to see all quotes with that tag.",
            "  • «Quote #» matches the number on the All Quotes sheet.",
            "",
            "Tips",
            "  • Click the drop-down arrows in the header to filter.",
            "  • Header row stays visible while you scroll.",
            f"  • Source website: {BASE_URL}",
            f"  • Quotes collected: {len(data)}",
        ]
        for i, line in enumerate(help_lines, start=1):
            cell = sheet_help.cell(row=i, column=1, value=line)
            if i == 1:
                cell.font = Font(bold=True, size=14, color="2F5496")
            elif line.startswith("Sheet") or line == "Tips":
                cell.font = Font(bold=True, size=11)
        sheet_help.column_dimensions["A"].width = 90

        workbook.save(output_path)
        print(
            f"[Success] Saved: {output_path.name} | "
            f"{len(data)} quotes (All Quotes), "
            f"{tag_rows} rows (Tags Lookup)."
        )
        return output_path

    except PermissionError:
        print(
            f"[Error] Cannot write '{output_path}'. "
            "Close the file in Excel and run again."
        )
    except OSError as error:
        print(f"[Error] Could not save file: {error}")
    except Exception as error:
        print(f"[Error] Excel error: {error}")

    return None


# ------------------------------------------------------------
# 10. CLI helpers
# ------------------------------------------------------------
def parse_page_argument(argv: list[str]) -> int | None:
    """
    Read page count from CLI.
    Returns None if the argument is invalid.
    """
    if len(argv) <= 1:
        return DEFAULT_PAGES

    raw = argv[1].strip()
    try:
        value = int(raw)
    except ValueError:
        print(f"[Error] '{raw}' is not a number.")
        print("Usage: python quotes_scraper.py [pages]")
        print("Example: python quotes_scraper.py 10")
        return None

    if value < 1:
        print("[Error] Page count must be >= 1.")
        return None

    if value > 100:
        print("[Warning] Very large page count. Capping at 100 for safety.")
        value = 100

    return value


# ------------------------------------------------------------
# 11. Main
# ------------------------------------------------------------
def main() -> None:
    """Entry point: scrape → preview → Excel."""
    max_pages = parse_page_argument(sys.argv)
    if max_pages is None:
        sys.exit(1)

    print("=" * 50)
    print("Quotes Scraper")
    print(f"Site:   {BASE_URL}")
    print(f"Pages:  {max_pages}")
    print(f"Delay:  {DELAY_SECONDS}s between requests")
    print("=" * 50)

    try:
        quotes = scrape_pages(max_pages)
    except KeyboardInterrupt:
        print("\n[Info] Interrupted by user.")
        sys.exit(130)
    except Exception as error:
        print(f"[Error] Unexpected failure during scraping: {error}")
        sys.exit(1)

    if not quotes:
        print("[Error] No data collected. Check your internet connection or try later.")
        sys.exit(1)

    print("\n--- Sample (first 3 quotes) ---")
    for item in quotes[:3]:
        text = item["quote_text"]
        preview = text if len(text) <= 70 else text[:70] + "..."
        print(f"  Quote:  {preview}")
        print(f"  Author: {item['author']}")
        print(f"  Tags:   {', '.join(item['tags']) if item['tags'] else '(none)'}")
        print()

    saved = save_to_excel(quotes, filename=OUTPUT_FILE)
    if saved is None:
        sys.exit(1)

    print("[Info] Done. Open the Excel file to view results.")
    print("[Info] Read README.txt for usage instructions.")


if __name__ == "__main__":
    main()
